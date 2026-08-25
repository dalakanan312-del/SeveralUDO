"""Build the tracker's bundled medieval-name library from Decades Names.

Usage:
    python tools/build_medieval_names.py SOURCE.xlsx app/medieval_names.json

Only the ``medieval names`` worksheet is read. The workbook's Boy, Girl, and
surname columns become the tracker's Male, Female, and Surname collections.
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path

import openpyxl


SHEET_NAME = "medieval names"
SOURCE_URL = (
    "https://docs.google.com/spreadsheets/d/"
    "1Cr-MFsjQycEF17XsVXZrjScwF8z39wmHJCLXlpBOWPU/edit?gid=969243054"
)
CULTURE_COLUMNS = {
    "English": 3,
    "Irish": 6,
    "Welsh": 9,
    "Scottish": 12,
    "Swedish": 15,
    "French": 18,
    "Greek": 21,
    "German": 24,
    "Russian": 27,
    "Italian": 30,
    "Spanish": 33,
    "Turkish": 36,
    "Korean": 39,
}
KINDS = ("Male", "Female", "Surname")


def unique_values(sheet, column: int) -> list[str]:
    values: list[str] = []
    seen: set[str] = set()
    for row in range(3, sheet.max_row + 1):
        raw = sheet.cell(row, column).value
        value = str(raw).strip() if raw not in (None, "") else ""
        key = value.casefold()
        if not value or key in seen:
            continue
        seen.add(key)
        values.append(value)
    return values


def build(source: Path, destination: Path) -> dict:
    workbook = openpyxl.load_workbook(source, data_only=True, read_only=False)
    sheet = workbook[SHEET_NAME]
    cultures = {
        culture: {
            kind: unique_values(sheet, first_column + offset)
            for offset, kind in enumerate(KINDS)
        }
        for culture, first_column in CULTURE_COLUMNS.items()
    }
    payload = {
        "schema_version": 1,
        "collection": "Medieval names",
        "source": {
            "title": "Decades Names",
            "sheet": SHEET_NAME,
            "url": SOURCE_URL,
            "spreadsheet_id": "1Cr-MFsjQycEF17XsVXZrjScwF8z39wmHJCLXlpBOWPU",
            "source_modified": "2026-02-04",
        },
        "cultures": cultures,
    }
    destination.parent.mkdir(parents=True, exist_ok=True)
    destination.write_text(
        json.dumps(payload, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )
    return payload


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("source", type=Path)
    parser.add_argument("destination", type=Path)
    args = parser.parse_args()
    payload = build(args.source, args.destination)
    counts = {
        culture: {kind: len(values) for kind, values in groups.items()}
        for culture, groups in payload["cultures"].items()
    }
    print(json.dumps(counts, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
